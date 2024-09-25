from fastapi.responses import FileResponse

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

import os


async def edit_file(filepath: str, rows: list):
    wb = load_workbook(filepath)
    ws = wb.active

    integer_style = NamedStyle(name="integer_style", number_format='0')

    for row in rows:
        for cell in ws[row]:  
            if isinstance(cell.value, (int, float)):
                cell.style = integer_style
    
    wb.save(filepath)

def get_unique_filename(directory, filename):
    base, extension = os.path.splitext(filename)
    counter = 1
    unique_filename = filename
    while os.path.exists(os.path.join(directory, unique_filename)):
        unique_filename = f"{base}_({counter}){extension}"
        counter += 1

    return unique_filename

async def get_shablon(directory):
    filename = "shablon.xlsx"
    shablon_location = os.path.join(directory, filename)
    return FileResponse(path=shablon_location, filename=filename)