
from fastapi import HTTPException, status

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.engine import Result
from sqlalchemy import select

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

from app.core.models import Parser, File
from app.core.config import settings

import pandas as pd


async def get_all_data_from_file(file_id: int, user_id: int, session: AsyncSession):
    stmt = (
        select(Parser).where(Parser.user_id == user_id).where(Parser.file_id == file_id)
    )
    result: Result = await session.execute(stmt)
    parsers_data = result.scalars().all()
    if parsers_data == []:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="У юзера нет такого файла"
        )

    return parsers_data


def edit_price(data):
    if data.best_price.lower() == "пусто":
        return None

    if data.nds.lower() == "нет":
        price_company = int(float(data.price) * 1.07)
        price_from_site = int(data.best_price)

        if price_company < price_from_site:
            best_price = price_from_site - 1
        else:
            best_price = price_company
    else:
        price_company = int(float(data.price) * 1.27)
        price_from_site = int(data.best_price)
        if price_company < price_from_site:
            best_price = price_from_site - 1
        else:
            best_price = price_company

    return best_price


async def get_file(file_id: int, session: AsyncSession, user_id: int):
    stmt = select(File).where(File.id == file_id).where(File.user_id == user_id)
    result: Result = await session.execute(stmt)
    file = result.scalar()

    if file is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="У юзера нет такого файла"
        )

    return file


async def get_filename(file_id: int, session: AsyncSession, user_id: int):
    file = await get_file(file_id=file_id, user_id=user_id, session=session)
    return file.after_parsing_filename


async def set_filename(file_id: int, session: AsyncSession, user_id: int):
    file = await get_file(file_id=file_id, user_id=user_id, session=session)

    if "обработанный_" in file.after_parsing_filename:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT, detail="уже обработан"
        )
    file.after_parsing_filename = f"обработанный_{file.after_parsing_filename}"
    session.add(file)
    await session.commit()

async def edit_file(filepath: str):
    wb = load_workbook(filepath)
    ws = wb.active

    integer_style = NamedStyle(name="integer_style", number_format='0')

    for row in ["J", "K", "L", "M", "N", "F"]:
        for cell in ws[row]:  
            if isinstance(cell.value, (int, float)):
                cell.style = integer_style
    
    wb.save(filepath)

async def to_file(filename: str, parser_data: list, session: AsyncSession):
    try:
        # Определяем базовые столбцы
        columns = [
            "Артикул",
            "Наименование",
            "Брэнд",
            "Артикул",
            "Кол-во",
            "Цена",
            "Партия",
            "НДС",
            "Лого",
            "Доставка",
            "Лучшая цена",
            "Количество",
        ]

        # Проверка и добавление дополнительных столбцов
        has_new_price = any(hasattr(data, 'new_price') for data in parser_data)
        has_after_vat_price = any(hasattr(data, 'after_vat_price') for data in parser_data)

        if has_new_price:
            columns.append("Цена с лого")
        if has_after_vat_price:
            columns.append("Цена после расчёта")

        # Формирование данных для Excel
        excel = [
            [
                data.article,
                data.name,
                data.brand,
                data.article1,
                data.quantity,
                float(data.price),
                data.batch,
                data.nds,
                data.logo,
                int(data.delivery_time),
                int(data.best_price),
                int(data.quantity1),
                *([int(data.new_price)] if hasattr(data, 'new_price') and has_new_price and data.new_price != "-1" else []),  # Добавляем 'new_price', если он существует
                *([int(data.after_vat_price)] if hasattr(data, 'after_vat_price') and has_after_vat_price else [])  # Добавляем 'after_vat_price', если он существует
            ]
            for data in parser_data
        ]
        if len(excel[0]) == 13:
            columns.remove("Цена с лого")
            for data in parser_data:
                data.new_price = None
                session.add(data)
                await session.commit()

        

        # Создание DataFrame
        df = pd.DataFrame(excel, columns=columns)

        # Сохранение в файл Excel
        output_path = f"{str(settings.upload.path_for_upload)}/обработанный_{filename}"
        df.to_excel(output_path, index=False)
        await edit_file(output_path)

    except AttributeError as e:
        print(f"Ошибка атрибутов в данных: {e}")
    except FileNotFoundError as e:
        print(f"Ошибка: путь не найден или недоступен: {e}")
    except PermissionError as e:
        print(f"Ошибка прав доступа к файлу: {e}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

async def get_all_files(session: AsyncSession, user_id: int):
    stmt = select(File).where(File.user_id == user_id)
    result: Result = await session.execute(stmt)

    return result.scalars().all()
