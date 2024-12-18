from fastapi import HTTPException, status
from fastapi.responses import FileResponse

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.engine import Result
from sqlalchemy import select

from app.core.models import File
from app.core.config import settings
from . import crud

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

import os
import pandas as pd


async def get_files(session: AsyncSession, user_id: int):
    stmt = select(File).where(File.user_id==user_id).order_by(File.date)
    result: Result = await session.execute(stmt)
    files = result.scalars().all()
    return files

def zero_files(files):
    if files == []:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Файлов нет у данного пользователя"
        )


def check_same(file):
    try:
        RIGHT_DATA = [i for i in pd.read_excel("app/upload_file/shablon.xlsx").values.tolist() if "Артикул" in i][0]
        file_upload_user = [list(map(str, row)) for row in pd.read_excel(file).values.tolist()]
        head = file_upload_user[3]
        rows_with_data = file_upload_user[4:]
        for en, i in enumerate(rows_with_data):
            if not (head == RIGHT_DATA and len(i) - i.count('nan') == len(RIGHT_DATA)):
                print(en, i)
        return head == RIGHT_DATA and all([len(i) - i.count('nan') == len(RIGHT_DATA) for i in rows_with_data])
    except:
        return False

def check_has_last_file_after_parsing(last_file):
    if not last_file.is_after_parsing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Последний файл еще не спаршен"
        )

async def edit_file(filepath: str, rows: list):
    wb = load_workbook(filepath)
    ws = wb.active

    integer_style = NamedStyle(name="integer_style", number_format='0')

    for row in rows:
        for cell in ws[row]:  
            if cell.value and cell.value.isdigit():
                cell.style = integer_style
    
    os.remove(filepath)
    wb.save(filepath)